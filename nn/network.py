import json
import os
import sys
import pandas as pd
import win32com.client
from nn.layer import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sklearn.model_selection import train_test_split
from tqdm import tqdm


class Network():
    wb: Workbook
    layers: list[Layer]
    random_state: int
    filename: str

    def __init__(
        self,
        csv_path: str,
        config_path: str,
        random_state: int,
    ) -> None:
        self.wb = load_workbook('template.xlsm', keep_vba=True)
        self.wb.active.title = 'Main'
        self.wb.create_sheet('Training Data')
        self.wb.create_sheet('Test Data')
        self.random_state = random_state
        self.filename = csv_path.split('/')[-1].replace('.csv', '')

        self.init_data(csv_path)
        self.init_layers(config_path)
        self.init_predictions()
        self.save()
        self.inject_macros()
        self.save()

    def init_data(self, csv_path: str):
        df = pd.read_csv(csv_path)
        self.input_size = df.shape[1] - 1

        X = df.iloc[:, :-1].apply(pd.to_numeric)
        y = df.iloc[:, -1].apply(pd.to_numeric)
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.3, random_state=self.random_state, stratify=y)

        train_sheet = self.wb['Training Data']
        train_sheet.append(list(df.columns))
        for i in range(len(X_train)):
            row = list(X_train.iloc[i])
            row.append(y_train.iloc[i])
            train_sheet.append(row)

        test_sheet = self.wb['Test Data']
        test_sheet.append(list(df.columns))
        for i in range(len(X_test)):
            row = list(X_test.iloc[i])
            row.append(y_test.iloc[i])
            test_sheet.append(row)

    def init_layers(self, config_path):
        self.configs = json.load(open(config_path))
        self.layers = []
        max_input_size = 0
        previous_layer_output_size = self.input_size
        for i, config in tqdm(enumerate(self.configs['layers']), desc='Initializing weights'):
            max_input_size = max(config['input_size'], max_input_size)
            if config['input_size'] != previous_layer_output_size:
                raise ValueError('Bad input_size/output_size.')
            previous_layer_output_size = config['output_size']

            sheet = self.wb.create_sheet(f'Layer_{i + 1}')
            layer = Layer(config)
            rows = [list(row) for row in layer.weights]
            for row in rows:
                sheet.append(list(row))
            self.layers.append(layer)

        self.wb.create_sheet('Bias')
        for i, layer in enumerate(self.layers):
            for j, bias in enumerate(layer.bias):
                self.wb['Bias'].cell(row=(j + 1), column=(i + 1)).value = bias

        for i in range(len(self.layers)):
            self.wb.create_sheet(f'Z_{i + 1}')
        for i in range(len(self.layers)):
            self.wb.create_sheet(f'A_{i + 1}')

    def init_predictions(self):
        train_predictions = self.wb.copy_worksheet(self.wb['Training Data'])
        train_predictions.title = 'Training Predictions'
        for i in range(1, train_predictions.max_column):
            column = get_column_letter(i)
            train_predictions.column_dimensions[column].hidden = True

        test_predictions = self.wb.copy_worksheet(self.wb['Test Data'])
        test_predictions.title = 'Test Predictions'
        for i in range(1, test_predictions.max_column):
            column = get_column_letter(i)
            test_predictions.column_dimensions[column].hidden = True

    def inject_macros(self):
        macros = []

        # global_variables = (f'Public nLayers As Integer\n'
        #                     f'Public Activation As Variant\n'
        #                     f'Public Alpha As Integer\n'
        #                     f'Public Epoch As Integer\n')
        # macros.append(global_variables)

        # # inject constants
        # epochs = self.configs.get('epochs', 10)
        # alpha = self.configs.get('alpha', 0.1)
        # constants = (f'Sub InitParams()\n'
        #              f'    Epochs = {epochs}\n'
        #              f'    Alpha = {alpha}\n'
        #              f'End Sub')
        # macros.append(constants)

        for file in os.listdir('macros'):
            with open(f'macros/{file}', 'r') as file:
                macros.append(file.read())

        _file = os.path.abspath(sys.argv[0])
        path = os.path.join(os.path.dirname(_file), f'{self.filename}.xlsm')
        excel = win32com.client.Dispatch("Excel.Application")

        try:
            excel.Visible = False
            excel.DisplayAlerts = False
            # need to enable access
            # https://stackoverflow.com/questions/25638344/programmatic-access-to-visual-basic-project-is-not-trusted
            wb = excel.Workbooks.Open(Filename=path)

            for macro in tqdm(macros, desc='Injecting macros'):
                excelModule = wb.VBProject.VBComponents.Add(1)
                excelModule.CodeModule.AddFromString(macro)
                wb.SaveAs(path)

            excel.Workbooks(1).Close(SaveChanges=1)

        except Exception as e:
            print(f'Unable to load {path}: {e}')

        finally:
            excel.Application.Quit()
            del excel

        self.wb = load_workbook(path, keep_vba=True)

    def save(self) -> None:
        path = f'{self.filename}.xlsm'
        self.wb.save(path)

        # weird behavior
        # https://stackoverflow.com/questions/59585265/is-there-any-way-to-create-a-xlsm-file-from-scratch-in-python
        self.wb = load_workbook(path, keep_vba=True)
        self.wb.save(path)
